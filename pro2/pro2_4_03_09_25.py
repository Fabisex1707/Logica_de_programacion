#read excel by Openpyxl

from tabulate import tabulate
import openpyxl


# libro=openpyxl.load_workbook('MiArchivoExcel.xlsx')
# print(type(libro))
# print(libro.sheetnames)
# hoja=libro["Sheet1"]
# #buesqueda por etiqueta o nombre
# valor_celda_A1=hoja["A1"].value
# valor_celda_B1=hoja["B1"].value
# print(valor_celda_A1)
# print(type(valor_celda_A1))

# print(valor_celda_A1)
# print(type(valor_celda_A1))

# #busqueda por cordenadas
# valor_celda_B1_cordenadas=hoja.cell(row=1,column=2).value
# print(valor_celda_B1_cordenadas)

# #acceso a un rango de celdas
# for row in range(1,8):
#     print(f"renglon {row}, {hoja.cell(row,column=2).value}")

# rango_celdas=hoja["A1":"C7"]

# for row in rango_celdas:
#     for celda in row:
#         print(f"Celda{celda.coordinate} = {celda.value}")
#     print("---Fin del renglon---\n")


datos_excel=openpyxl.load_workbook('MiArchivoExcel.xlsx')
data_frame=datos_excel.active

data=[]

for row in range(0,data_frame.max_row):
     _row=[row+1,]

     for colum in data_frame.iter_cols(1,data_frame.max_column):#comienza en 0 por devolver una tupla
        _row.append(colum[row].value)
     data.append(_row)
     
headers=["#","Hora","Fruta","Cantidad"]
print(tabulate(data,headers=headers,tablefmt="fancy_grid"))





